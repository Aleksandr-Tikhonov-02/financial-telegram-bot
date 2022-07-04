import notifiers
from notifiers import get_notifier
from keys import *
import finplot as fplt
import yfinance
import yfinance as yf
import telebot
import finnhub
import openpyxl
from telebot import types

bot = telebot.TeleBot(token)
finnhub_client = finnhub.Client(api_key=finhub_api_key)
keyboard = types.InlineKeyboardMarkup()
notifier = notifiers.get_notifier('telegram')

name_of_stock = ''
amount_of_stocks = 0

title = ''
amount = 0
price = 0

name_column = 1
amount_column = 2
price_column = 3

list_of_symbols = []


def init(id):
    wb = openpyxl.load_workbook('usr.xlsx')
    wb.create_sheet(title=str(id))
    sh = wb[str(id)]
    current_row_cell = sh.cell(row=1, column=1)
    current_row_cell.value = 2
    wb.save('usr.xlsx')


@bot.message_handler(commands=['start', 'help'])
def start(message):
    init(message.from_user.id)
    start_message = 'Привет! Это финансовый бот, который поможет тебе при работе с акциями!\n\n' + \
                    'Команды:\n\n' + \
                    '/portfolio - чтобы посмотреть твой портфель\n' + \
                    '/quotations - чтобы посмотреть котировки \n' + \
                    '/graphics - чтобы посмотреть графики\n' + \
                    '/notifications - чтобы посмотреть увледомления или установить новые \n'
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    key_portfolio = types.KeyboardButton(text='Портфель')
    key_quotations = types.KeyboardButton(text='Котировки')
    key_graphics = types.KeyboardButton(text='Графики')
    key_notif = types.KeyboardButton(text='Увледомления')
    markup.add(key_portfolio, key_graphics, key_quotations, key_notif)
    bot.send_message(message.chat.id, start_message, reply_markup=markup)


@bot.callback_query_handler(func=lambda call: True)
def callback_worker(call):
    if call.data == 'hour' or 'day' or 'week' or 'month':
        get_graphics(call.message, call.data)


@bot.message_handler(content_types=['text'])
def random_answers(message):
    if message.text == 'Портфель':
        portfolio(message)
    if message.text == 'Графики':
        graphics(message)
    if message.text == 'Котировки':
        quotations(message)
    if message.text == 'Увледомления':
        notifications(message)
    if message.text == 'Показать акции':
        show_stocks(message)
    if message.text == 'Купить акции':
        add_stock(message)
    if message.text == 'Продать акции':
        sell_stocks(message)
    if message.text == 'Вернуться в меню':
        start(message)


@bot.message_handler(commands=['portfolio'])
def portfolio(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    key_show_stocks = types.KeyboardButton(text='Показать акции')
    key_new_stock = types.KeyboardButton(text='Купить акции')
    key_sell_stock = types.KeyboardButton(text='Продать акции')
    key_back_to_menu = types.KeyboardButton(text='Вернуться в меню')
    markup.add(key_show_stocks, key_new_stock, key_sell_stock, key_back_to_menu)
    bot.send_message(message.chat.id, "Выберите команду", reply_markup=markup)


# ПОРТФОЛИО
def show_stocks(message):
    global current_row
    global name_column
    global price_column
    wb = openpyxl.load_workbook('usr.xlsx')
    sh = wb[str(message.from_user.id)]
    current_row_cell = sh.cell(row=1, column=1)
    current_row = current_row_cell.value
    if current_row == 2:
        answer = '💼 Портфель\n\nПусто'
        bot.send_message(message.chat.id, answer)
    else:
        answer = '💼 Портфель💰💰💰 \n\n'
        for i in range(2, sh.max_row + 1):
            answer += str(i - 1) + ". "
            n = sh.cell(row=i, column=name_column)
            a = sh.cell(row=i, column=amount_column)
            p = sh.cell(row=i, column=price_column)
            answer += (str(n.value) + ': ').ljust(6) + (str(a.value) + ' шт. цена: ').ljust(12) + str(p.value) + '\n'
        answer += '\n'
        bot.send_message(message.chat.id, answer)


def add_stock(message):
    bot.send_message(message.chat.id, 'Введи название акции')
    bot.register_next_step_handler(message, get_title_buy)


def sell_stocks(message):
    bot.send_message(message.chat.id, 'Введи название акции')
    bot.register_next_step_handler(message, get_title_sell)
    # проверка на существуемость: если нет то ошибка и еще раз


# получение даннхы для покупки
def get_title_buy(message):
    stocks_variants = finnhub_client.symbol_lookup(message.text)
    if stocks_variants['count'] == 0:
        bot.send_message(message.chat.id, 'Не найдено подходящих акций')
    else:
        answer = 'Мы нашли похожие акции, выберите одну из них:\n'
        for stockEx in stocks_variants["result"]:
            answer += ('• ' + stockEx["symbol"] + '(' + stockEx["description"] + ', ' + stockEx["type"] + ')' + '\n')
            list_of_symbols.append(stockEx["symbol"])
        bot.send_message(message.chat.id, answer)
        bot.register_next_step_handler(message, get_title_buy_next)


def get_title_buy_next(message):
    global title
    if message.text not in list_of_symbols:
        bot.send_message(message.chat.id, 'Нет такой акции в верхнем списке')
    else:
        title = message.text
        list_of_symbols.clear()
        bot.send_message(message.chat.id, 'Введи количество акций')
        bot.register_next_step_handler(message, get_amount_buy)


def get_amount_buy(message):
    global amount
    amount = int(message.text)
    bot.send_message(message.chat.id, 'Введи цену покупки')
    bot.register_next_step_handler(message, get_price_buy)


def get_price_buy(message):
    global price
    price = int(message.text)
    buy(message)


# получение даннхы для продажи
def get_title_sell(message):
    global title
    title = message.text
    bot.send_message(message.chat.id, 'Введи количество акций')
    bot.register_next_step_handler(message, get_amount_sell)


def get_amount_sell(message):
    global amount
    amount = int(message.text)
    bot.send_message(message.chat.id, 'Введи цену продажи')
    bot.register_next_step_handler(message, get_price_sell)


def get_price_sell(message):
    global price
    price = int(message.text)
    sell(message)


# покупка и продажа под капотом
def buy(message):
    global title
    global amount
    global price

    global name_column
    global price_column
    global amount_column

    if amount <= 0 or price <= 0:
        answer = '💼 Дорогой, инвестор\n\n '
        answer += ' Кол-во акций и цена должны быть положительными. '
        bot.send_message(message.chat.id, answer)
        show_stocks(message)
        return

    wb = openpyxl.load_workbook('usr.xlsx')
    sh = wb[str(message.from_user.id)]

    current_row_cell = sh.cell(row=1, column=1)
    current_row = current_row_cell.value

    alreadyExist, row_value_existed = is_existed(sh)

    if not alreadyExist:
        current_row_cell = sh.cell(row=1, column=1)
        t = sh.cell(row=current_row, column=name_column)
        a = sh.cell(row=current_row, column=amount_column)
        p = sh.cell(row=current_row, column=price_column)
        t.value = title
        a.value = amount
        p.value = price
        current_row_cell.value = current_row + 1
    else:
        a = sh.cell(row=row_value_existed, column=amount_column)
        p = sh.cell(row=row_value_existed, column=price_column)
        tmp = int(a.value)
        a.value = tmp + int(amount)
        p.value = (int(p.value) * tmp + int(price) * int(amount)) / (tmp + int(amount))

    wb.save('usr.xlsx')
    title = ''
    amount = 0
    price = 0
    show_stocks(message)


def sell(message):
    global title
    global amount
    global price

    global name_column
    global price_column
    global amount_column

    if amount <= 0 or price <= 0:
        answer = '💼 Дорогой, инвестор\n\n '
        answer += ' Кол-во акций и цена должны быть положительными. '
        bot.send_message(message.chat.id, answer)
        show_stocks(message)
        return

    wb = openpyxl.load_workbook('usr.xlsx')
    sh = wb[str(message.from_user.id)]
    alreadyExist, row_value_existed = is_existed(sh)

    if alreadyExist:
        a = sh.cell(row=row_value_existed, column=amount_column)
        p = sh.cell(row=row_value_existed, column=price_column)
        t = sh.cell(row=row_value_existed, column=name_column)
        tmp = a.value
        if a.value < int(amount):
            answer = '💼 Дорогой, инвестор\n\n '
            answer += ' Кол-во акций не может быть меньше, чем Вы можете продать. '
            bot.send_message(message.chat.id, answer)
            show_stocks(message)
            return
        a.value = tmp - int(amount)
        if a.value == 0:
            shift_of_column(message.from_user.id, row_value_existed)
        else:
            wb.save('usr.xlsx')

        answer = '💼 Время - деньги\n\n'
        answer += 'Вы продали акции ' + t.value + ' в кол-ве ' + str(amount) + ' \n\n'
        answer += 'Прибыль: ' + str(int(price) * int(amount) - int(p.value) * int(amount))
        bot.send_message(message.chat.id, answer)
    else:
        answer = '💼 В портфели\n\nНет акций с титром ' + title
        bot.send_message(message.chat.id, answer)

    title = ''
    amount = 0
    price = 0


def is_existed(sh):
    current_row_cell = sh.cell(row=1, column=1)
    max_row = current_row_cell.value
    alreadyExist = False

    cntr = 1
    for row in sh.iter_rows(max_row=max_row):
        for cell in row:
            if str(cell.value) == title:
                alreadyExist = True
                break
        if alreadyExist:
            break
        cntr += 1
    return alreadyExist, cntr


def shift_of_column(id_user, id_row_of_null):
    wb = openpyxl.load_workbook('usr.xlsx')
    sh = wb[str(id_user)]
    if id_row_of_null == sh.max_row:
        n = sh.cell(row=id_row_of_null, column=name_column)
        a = sh.cell(row=id_row_of_null, column=amount_column)
        p = sh.cell(row=id_row_of_null, column=price_column)
        n.value = None
        a.value = None
        p.value = None
        current_row_cell = sh.cell(row=1, column=1)
        current_row_cell.value = int(current_row_cell.value) - 1
        wb.save('usr.xlsx')
        return

    n = sh.cell(row=id_row_of_null + 1, column=name_column)
    a = sh.cell(row=id_row_of_null + 1, column=amount_column)
    p = sh.cell(row=id_row_of_null + 1, column=price_column)

    column_of_name = [n.value]
    column_of_amount = [a.value]
    column_of_price = [p.value]

    for i in range(id_row_of_null + 2, sh.max_row + 1):
        n = sh.cell(row=i, column=name_column)
        a = sh.cell(row=i, column=amount_column)
        p = sh.cell(row=i, column=price_column)

        column_of_name.append(n.value)
        column_of_amount.append(a.value)
        column_of_price.append(p.value)

    column_of_name.reverse()
    column_of_amount.reverse()
    column_of_price.reverse()

    for i in range(id_row_of_null, sh.max_row):
        n = sh.cell(row=i, column=name_column)
        a = sh.cell(row=i, column=amount_column)
        p = sh.cell(row=i, column=price_column)
        n.value = column_of_name.pop()
        a.value = column_of_amount.pop()
        p.value = column_of_price.pop()

    n = sh.cell(row=sh.max_row, column=name_column)
    a = sh.cell(row=sh.max_row, column=amount_column)
    p = sh.cell(row=sh.max_row, column=price_column)
    n.value = None
    a.value = None
    p.value = None

    current_row_cell = sh.cell(row=1, column=1)
    current_row_cell.value = int(current_row_cell.value) - 1
    wb.save('usr.xlsx')


@bot.message_handler(commands=['graphics'])
def graphics(message):
    bot.send_message(message.chat.id, 'Введите название акции')
    bot.register_next_step_handler(message, get_title_for_graphics)


# ГРАФИКИ
def get_title_for_graphics(message):
    global name_of_stock
    name_of_stock = message.text
    stocks_variants = finnhub_client.symbol_lookup(name_of_stock)
    if stocks_variants['count'] == 0:
        bot.send_message(message.chat.id, 'Не найдено подходящих акций')
    else:
        answer = 'Мы нашли похожие акции, выберите одну из них:\n'
        for stockEx in stocks_variants["result"]:
            answer += ('• ' + stockEx["symbol"] + '(' + stockEx["description"] + ', ' + stockEx["type"] + ')' + '\n')
            list_of_symbols.append(stockEx["symbol"])
        bot.send_message(message.chat.id, answer)
        bot.register_next_step_handler(message, get_title_for_graphics_next)


def get_title_for_graphics_next(message):
    global list_of_symbols
    global name_of_stock
    if message.text not in list_of_symbols:
        bot.send_message(message.chat.id, 'Нет такой акции в верхнем списке')
        # bot.register_next_step_handler(message, get_title_of_stock_for_quotations_next)
    else:
        name_of_stock = message.text
        markup = types.InlineKeyboardMarkup()
        key_hour = types.InlineKeyboardButton(text='Час', callback_data='hour')
        key_day = types.InlineKeyboardButton(text='День', callback_data='day')
        key_week = types.InlineKeyboardButton(text='Неделя', callback_data='week')
        key_month = types.InlineKeyboardButton(text='Месяц', callback_data='month')
        markup.add(key_hour, key_day, key_week, key_month)
        bot.send_message(message.chat.id, 'Выберите период времени', reply_markup=markup)


def get_graphics(message, time_choice):
    df = None
    if time_choice == 'hour':
        df = yf.download(tickers=name_of_stock, period="1h", interval="5m")
    if time_choice == 'day':
        df = yf.download(tickers=name_of_stock, period="1d", interval="1h")
    if time_choice == 'week':
        df = yf.download(tickers=name_of_stock, period="5d", interval="1h")
    if time_choice == 'month':
        df = yf.download(tickers=name_of_stock, period="1mo", interval="1d")
    try:
        fplt.candlestick_ochl(df[['Open', 'Close', 'High', 'Low']])
        stock = yfinance.Ticker(name_of_stock)
        bot.send_message(message.chat.id, 'Поиск информации....')
        fplt.timer_callback(save, 0.1, single_shot=True)
        fplt.timer_callback(close, 1, single_shot=False)
        fplt.show()
        bot.send_message(message.chat.id, 'Валюта: ' + str(stock.info["currency"]) + '\n')
        bot.send_photo(message.chat.id, photo=open('screenshot.png', 'rb'))
    except:
        fplt.timer_callback(close, 0.5, single_shot=False)
        bot.send_message(message.chat.id, 'Данная информация не доступна')


def save():
    fplt.screenshot(open('screenshot.png', 'wb'))


def close():
    fplt.close()


# КОТИРОВКИ
@bot.message_handler(commands=['quotations'])
def quotations(message):
    bot.send_message(message.chat.id, 'Введите название акции')
    bot.register_next_step_handler(message, get_title_of_stock_for_quotations)


def get_title_of_stock_for_quotations(message):
    global list_of_symbols
    stocks_variants = finnhub_client.symbol_lookup(message.text)
    if stocks_variants['count'] == 0:
        bot.send_message(message.chat.id, 'Не найдено подходящих акций')
    else:
        answer = 'Мы нашли похожие акции, выберите одну из них:\n'
        for stockEx in stocks_variants["result"]:
            answer += ('• ' + stockEx["symbol"] + '(' + stockEx["description"] + ', ' + stockEx["type"] + ')' + '\n')
            list_of_symbols.append(stockEx["symbol"])
        bot.send_message(message.chat.id, answer)
        bot.register_next_step_handler(message, get_title_of_stock_for_quotations_next)


def get_title_of_stock_for_quotations_next(message):
    global list_of_symbols
    if message.text not in list_of_symbols:
        bot.send_message(message.chat.id, 'Нет такой акции в верхнем списке')
        # bot.register_next_step_handler(message, get_title_of_stock_for_quotations_next)
    else:
        stock = yfinance.Ticker(message.text)
        answer = 'Текущая цена: ' + str(stock.info["currentPrice"]) + ' ' + str(stock.info["currency"]) + '\n'
        if stock.info["ask"] == 0:
            answer += 'Нельзя купить и продать'
        else:
            answer += 'Покупка: ' + str(stock.info["ask"]) + ' ' + str(stock.info["currency"]) + '\n' + \
                      'Продажа: ' + str(stock.info["bid"]) + ' ' + str(stock.info["currency"]) + '\n'

        bot.send_message(message.chat.id, answer)
        list_of_symbols.clear()


# УВЛЕДОМЛЕНИЯ
@bot.message_handler(commands=['notifications'])
def notifications(message):
    bot.send_message(message.chat.id, 'Введите название акции')
    bot.register_next_step_handler(message, get_title_of_stock_for_notifications)

def get_title_of_stock_for_notifications(message):
    global list_of_symbols
    stocks_variants = finnhub_client.symbol_lookup(message.text)
    if stocks_variants['count'] == 0:
        bot.send_message(message.chat.id, 'Не найдено подходящих акций')
    else:
        answer = 'Мы нашли похожие акции, выберите одну из них:\n'
        for stockEx in stocks_variants["result"]:
            answer += ('• ' + stockEx["symbol"] + '(' + stockEx["description"] + ', ' + stockEx["type"] + ')' + '\n')
            list_of_symbols.append(stockEx["symbol"])
        bot.send_message(message.chat.id, answer)
        bot.register_next_step_handler(message, get_title_of_stock_for_notifications_next)


def get_title_of_stock_for_notifications_next(message):
    global list_of_symbols
    if message.text not in list_of_symbols:
        bot.send_message(message.chat.id, 'Нет такой акции в верхнем списке')
        #  bot.register_next_step_handler(message, get_title_of_stock_for_notifications_next)
    else:
        pass


bot.polling(none_stop=True)
